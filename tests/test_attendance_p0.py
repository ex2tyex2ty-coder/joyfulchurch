from __future__ import annotations

import ast
import os
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from contextlib import closing
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook


APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from db import canonical_service_key, canonical_service_type, connect, init_db
from migration import migrate, optional_count, workbook_role


def _write_lineup_fixture(path: Path) -> None:
    """Create the smallest workbook that exercises attendance and lineup import."""
    workbook = Workbook()
    attendance = workbook.active
    attendance.title = "2099 예배인원"
    attendance.append(["날짜", "예배구분", "온라인", "현장", "합계"])
    attendance.append([date(2099, 1, 4), "주일예배", None, None, None])
    attendance.append([date(2099, 1, 11), "주일예배", 0, 0, 0])
    attendance.append([date(2099, 1, 18), "주일예배", 9, 101, 999])

    lineup = workbook.create_sheet("2099_1월")
    lineup.append(["", "주일", "주일", "주일"])
    lineup.append(["", date(2099, 1, 4), date(2099, 1, 11), date(2099, 1, 18)])
    lineup.append(["대표기도", "기도자1", "기도자2", "기도자3"])
    lineup.append(["FD", "담당자1", "담당자2", "담당자3"])
    lineup.append(["특별순서", "", "", "기존 특별순서"])
    workbook.save(path)


class AttendanceP0SchemaTest(unittest.TestCase):
    def test_init_db_adds_status_and_raw_source_columns(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "schema.db"
            init_db(db_path)
            with closing(connect(db_path)) as conn:
                columns = {
                    item["name"]: item
                    for item in conn.execute("PRAGMA table_info(attendance)").fetchall()
                }

            for name in (
                "record_status",
                "raw_online_count",
                "raw_offline_count",
                "raw_total_count",
                "metric_type",
                "measurement_note",
            ):
                self.assertIn(name, columns)
            self.assertIn("UNKNOWN", str(columns["record_status"]["dflt_value"]))

    def test_init_db_upgrades_an_existing_attendance_table_without_data_loss(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "legacy.db"
            init_db(db_path)
            with closing(sqlite3.connect(db_path)) as conn:
                conn.execute(
                    "INSERT INTO attendance("
                    "service_date,service_type,online_count,offline_count,total_count,data_quality) "
                    "VALUES('2099-01-04','주일예배',9,101,110,'Imported')"
                )
                conn.commit()

            init_db(db_path)

            with closing(connect(db_path)) as conn:
                saved = conn.execute(
                    "SELECT service_date,online_count,offline_count,total_count,record_status "
                    "FROM attendance WHERE id=1"
                ).fetchone()
            self.assertEqual(saved["service_date"], "2099-01-04")
            self.assertEqual(saved["total_count"], 110)
            self.assertIn(saved["record_status"], {"COUNTED", "UNKNOWN"})


class AttendanceP0ImportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.source_dir = Path(self.temp.name) / "source"
        self.source_dir.mkdir()
        self.db_path = Path(self.temp.name) / "attendance.db"
        self.report_path = Path(self.temp.name) / "report.json"
        _write_lineup_fixture(self.source_dir / "lineup.xlsx")

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _migrate(self) -> None:
        migrate(
            source_dir=self.source_dir,
            db_path=self.db_path,
            reset=False,
            report_path=self.report_path,
        )

    def test_attendance_and_lineup_share_one_canonical_service(self) -> None:
        self._migrate()
        with closing(connect(self.db_path)) as conn:
            for service_date in ("2099-01-04", "2099-01-11", "2099-01-18"):
                services = conn.execute(
                    "SELECT id,canonical_key FROM services "
                    "WHERE service_date=? AND service_type='주일예배'",
                    (service_date,),
                ).fetchall()
                attendance = conn.execute(
                    "SELECT service_id FROM attendance WHERE service_date=?",
                    (service_date,),
                ).fetchone()
                assignment_service_ids = {
                    item["service_id"]
                    for item in conn.execute(
                        "SELECT assignments.service_id FROM assignments "
                        "JOIN services ON services.id=assignments.service_id "
                        "WHERE services.service_date=? AND services.service_type='주일예배'",
                        (service_date,),
                    ).fetchall()
                }

                self.assertEqual(len(services), 1)
                self.assertTrue(services[0]["canonical_key"])
                self.assertEqual(assignment_service_ids, {attendance["service_id"]})

    def test_blank_and_explicit_zero_have_different_statuses_and_raw_values(self) -> None:
        self._migrate()
        with closing(connect(self.db_path)) as conn:
            blank = conn.execute(
                "SELECT * FROM attendance WHERE service_date='2099-01-04'"
            ).fetchone()
            explicit_zero = conn.execute(
                "SELECT * FROM attendance WHERE service_date='2099-01-11'"
            ).fetchone()

        self.assertEqual(blank["record_status"], "PENDING")
        self.assertIsNone(blank["raw_online_count"])
        self.assertIsNone(blank["raw_offline_count"])
        self.assertIsNone(blank["raw_total_count"])
        self.assertEqual(explicit_zero["record_status"], "UNKNOWN")
        self.assertEqual(explicit_zero["raw_online_count"], 0)
        self.assertEqual(explicit_zero["raw_offline_count"], 0)
        self.assertEqual(explicit_zero["raw_total_count"], 0)

    def test_original_total_is_preserved_while_normalized_total_is_recalculated(self) -> None:
        self._migrate()
        with closing(connect(self.db_path)) as conn:
            saved = conn.execute(
                "SELECT * FROM attendance WHERE service_date='2099-01-18'"
            ).fetchone()

        self.assertEqual(saved["raw_total_count"], 999)
        self.assertEqual(saved["online_count"], 9)
        self.assertEqual(saved["offline_count"], 101)
        self.assertEqual(saved["total_count"], 110)
        self.assertEqual(saved["record_status"], "COUNTED")
        self.assertEqual(saved["data_quality"], "Needs Review")

    def test_existing_event_is_relinked_to_the_canonical_service_and_attendance(self) -> None:
        init_db(self.db_path)
        with closing(connect(self.db_path)) as conn:
            event_id = conn.execute(
                "INSERT INTO events(title,series_key,category,event_date,status) "
                "VALUES('기존 특별순서','기존 특별순서','특별순서','2099-01-18','PLANNING')"
            ).lastrowid
            conn.commit()

        self._migrate()

        with closing(connect(self.db_path)) as conn:
            event = conn.execute("SELECT * FROM events WHERE id=?", (event_id,)).fetchone()
            attendance = conn.execute(
                "SELECT * FROM attendance WHERE service_date='2099-01-18'"
            ).fetchone()
        self.assertIsNotNone(event["service_id"])
        self.assertEqual(event["service_type"], "주일예배")
        self.assertEqual(attendance["service_id"], event["service_id"])
        self.assertEqual(attendance["event_id"], event_id)


class AttendanceP0InputContractTest(unittest.TestCase):
    def test_optional_count_accepts_only_non_negative_whole_counts(self) -> None:
        accepted = {
            95: 95,
            95.0: 95,
            "95.0": 95,
            "1,234": 1234,
        }
        for raw, expected in accepted.items():
            with self.subTest(raw=raw):
                self.assertEqual(optional_count(raw), expected)

        for raw in (95.5, -1, True, "nan", "inf"):
            with self.subTest(raw=raw):
                with self.assertRaises((TypeError, ValueError)):
                    optional_count(raw)

    def test_invalid_imported_count_is_unknown_and_needs_review(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source_dir = Path(directory) / "source"
            source_dir.mkdir()
            db_path = Path(directory) / "invalid.db"
            workbook = Workbook()
            attendance = workbook.active
            attendance.title = "2027 예배인원"
            attendance.append(["날짜", "예배구분", "온라인", "현장", "합계"])
            attendance.append([date(2027, 1, 3), "주일예배", True, 95.5, -1])
            lineup = workbook.create_sheet("2027_1월")
            lineup.append(["", "주일", "주일"])
            lineup.append(["", date(2027, 1, 3), date(2027, 1, 10)])
            workbook.save(source_dir / "invalid.xlsx")

            migrate(source_dir, db_path, reset=False, report_path=None)
            with closing(connect(db_path)) as conn:
                saved = conn.execute(
                    "SELECT record_status,data_quality,notes FROM attendance"
                ).fetchone()

        self.assertEqual(saved["record_status"], "UNKNOWN")
        self.assertEqual(saved["data_quality"], "Needs Review")
        self.assertIn("숫자로 읽을 수 없는", saved["notes"])

    def test_canonical_key_ignores_internal_service_type_spacing(self) -> None:
        spaced = canonical_service_key("2027-01-03", "주일 예배")
        compact = canonical_service_key("2027-01-03", "주일예배")

        self.assertEqual(spaced, compact)
        self.assertEqual(canonical_service_type(" 주일  예배 "), "주일 예배")

    def test_workbook_role_accepts_any_year_attendance_sheet_with_monthly_lineup(self) -> None:
        for year in (2026, 2027, 2099):
            with self.subTest(year=year):
                self.assertEqual(
                    workbook_role([f"{year} 예배인원", f"{year}_1월"]),
                    "LINEUP_ATTENDANCE",
                )

        self.assertEqual(workbook_role(["2027 예배인원"]), "UNKNOWN")
        self.assertEqual(workbook_role(["2027_1월"]), "UNKNOWN")
        self.assertEqual(workbook_role(["2027 예배인원", "2026_12월"]), "UNKNOWN")


class IsolatedAppMixin:
    @staticmethod
    def copy_isolated_project(sandbox: Path) -> Path:
        """Copy executable app files so imported defaults also point at temp/data."""
        for source in APP_DIR.glob("*.py"):
            shutil.copy2(source, sandbox / source.name)
        bible = APP_DIR / "bible_text.txt"
        if bible.exists():
            shutil.copy2(bible, sandbox / bible.name)
        return sandbox / "app.py"

    @staticmethod
    def load_helper_namespace() -> dict[str, object]:
        """Compile only pure attendance helpers; never execute app module/bootstrap."""
        tree = ast.parse((APP_DIR / "app.py").read_text(encoding="utf-8"))
        helper_names = {
            "_attendance_frame",
            "_last_elapsed_sunday",
            "_sunday_attendance_snapshot",
            "_scheduled_sunday_trend",
        }
        helper_nodes = [
            node for node in tree.body
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name in helper_names
        ]
        namespace: dict[str, object] = {
            "pd": pd,
            "date": date,
            "timedelta": __import__("datetime").timedelta,
            "ATTENDANCE_COUNTED_STATUSES": {"COUNTED", "ESTIMATED", "NO_STREAM"},
            "ATTENDANCE_CANCELLED_STATUSES": {"CANCELLED"},
            "ATTENDANCE_STATUS_LABELS": {
                "COUNTED": "집계 완료", "PENDING": "미입력", "CANCELLED": "예배 취소",
                "NO_STREAM": "온라인 송출 없음", "ESTIMATED": "추정 집계", "UNKNOWN": "확인 필요",
            },
        }
        exec(compile(ast.Module(body=helper_nodes, type_ignores=[]), str(APP_DIR / "app.py"), "exec"), namespace)
        return namespace


class AttendanceP0UiHelperTest(IsolatedAppMixin, unittest.TestCase):
    def setUp(self) -> None:
        namespace = self.load_helper_namespace()
        self.last_elapsed_sunday = namespace["_last_elapsed_sunday"]
        self.scheduled_sunday_trend = namespace["_scheduled_sunday_trend"]
        self.sunday_attendance_snapshot = namespace["_sunday_attendance_snapshot"]

    @staticmethod
    def _frame(records: list[dict[str, object]]) -> pd.DataFrame:
        data = pd.DataFrame(records)
        data["service_date"] = pd.to_datetime(data["service_date"])
        data["_record_status"] = data["record_status"]
        data["_counted"] = data["record_status"].isin({"COUNTED", "ESTIMATED", "NO_STREAM"})
        data["_cancelled"] = data["record_status"].isin({"CANCELLED"})
        return data

    def test_last_elapsed_sunday_excludes_today_until_the_day_has_elapsed(self) -> None:
        self.assertEqual(self.last_elapsed_sunday(date(2026, 8, 9)), date(2026, 8, 2))
        self.assertEqual(self.last_elapsed_sunday(date(2026, 8, 13)), date(2026, 8, 9))

    def test_snapshot_distinguishes_pending_row_from_absent_sunday(self) -> None:
        data = self._frame([
            {
                "service_date": "2026-07-19",
                "service_type": "주일예배",
                "offline_count": 123,
                "online_count": 9,
                "total_count": 132,
                "record_status": "COUNTED",
            },
            {
                "service_date": "2026-07-26",
                "service_type": "주일예배",
                "offline_count": None,
                "online_count": None,
                "total_count": None,
                "record_status": "PENDING",
            },
        ])

        snapshot = self.sunday_attendance_snapshot(data)

        self.assertEqual(snapshot["recorded_pending_dates"], [date(2026, 7, 26)])
        self.assertIn(date(2026, 8, 2), snapshot["absent_dates"])
        self.assertIn(date(2026, 8, 9), snapshot["absent_dates"])

    def test_scheduled_trend_keeps_a_gap_for_a_sunday_without_a_row(self) -> None:
        data = self._frame([
            {
                "service_date": "2026-07-26",
                "service_type": "주일예배",
                "offline_count": 100,
                "online_count": 10,
                "total_count": 110,
                "record_status": "COUNTED",
            },
            {
                "service_date": "2026-08-09",
                "service_type": "주일예배",
                "offline_count": 120,
                "online_count": 12,
                "total_count": 132,
                "record_status": "COUNTED",
            },
        ])

        trend = self.scheduled_sunday_trend(data, count_limit=3)
        missing = trend[trend["service_date"] == pd.Timestamp("2026-08-02")].iloc[0]

        self.assertEqual(missing["record_status_label"], "행 없음")
        self.assertTrue(pd.isna(missing["offline_count"]))


class AttendanceP0PageSmokeTest(IsolatedAppMixin, unittest.TestCase):
    def test_dashboard_and_attendance_page_render_the_reliability_summary(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            sandbox = Path(directory)
            isolated_app = self.copy_isolated_project(sandbox)
            runtime_db = sandbox / "data" / "joyful_worship_ops.db"
            runtime_db.parent.mkdir(parents=True, exist_ok=True)
            init_db(runtime_db)
            with closing(connect(runtime_db)) as conn:
                rows_to_insert = [
                    ("2026-07-19", 9, 123, 132, "COUNTED"),
                    ("2026-07-26", None, None, None, "PENDING"),
                ]
                for service_date, online, offline, total, status in rows_to_insert:
                    conn.execute(
                        "INSERT INTO attendance("
                        "service_date,service_type,online_count,offline_count,total_count,record_status,data_quality) "
                        "VALUES(?,?,?,?,?,?,?)",
                        (service_date, "주일예배", online, offline, total, status, "Imported"),
                    )
                conn.commit()
            smoke_script = """
from pathlib import Path
from streamlit.testing.v1 import AppTest
app = AppTest.from_file(str(Path('app.py').resolve()), default_timeout=30).run()
assert not app.exception, app.exception
pages = ['대시보드','예배 인원 현황']
for page in pages:
    app.sidebar.radio[0].set_value(page)
    app.run(timeout=30)
    assert not app.exception, (page, app.exception)
"""
            environment = os.environ.copy()
            environment["PYTHONPATH"] = str(sandbox)
            completed = subprocess.run(
                [sys.executable, "-c", smoke_script],
                cwd=sandbox,
                env=environment,
                capture_output=True,
                text=True,
                timeout=60,
            )
            self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)


if __name__ == "__main__":
    unittest.main(verbosity=2)
