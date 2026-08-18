from __future__ import annotations

import os
import shutil
import sqlite3
import tempfile
import threading
import urllib.request
import uuid
import zipfile
from contextlib import closing
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import DB_PATH, GOOGLE_SHEETS, GOOGLE_SHEETS_CACHE_DIR, ensure_directories
from db import connect, set_app_meta
from migration import migrate, workbook_role
from time_utils import iso_now_kst


SYNC_LOCK = threading.Lock()


def _download_xlsx(
    spreadsheet_id: str,
    destination: Path,
    timeout: int = 120,
    service_account_info: dict[str, Any] | None = None,
) -> None:
    if service_account_info:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload

        credentials = Credentials.from_service_account_info(
            service_account_info,
            scopes=["https://www.googleapis.com/auth/drive.readonly"],
        )
        drive = build("drive", "v3", credentials=credentials, cache_discovery=False)
        request = drive.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with destination.open("wb") as handle:
            downloader = MediaIoBaseDownload(handle, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    request = urllib.request.Request(url, headers={"User-Agent": "JOYFUL-WORSHIP-OPS/1.0"})
    with urllib.request.urlopen(request, timeout=timeout) as response, destination.open("wb") as handle:
        if response.status != 200:
            raise RuntimeError(f"Google Sheets 응답 오류: HTTP {response.status}")
        while chunk := response.read(1024 * 1024):
            handle.write(chunk)


def _validate_workbook(path: Path) -> dict[str, Any]:
    if not zipfile.is_zipfile(path):
        raise RuntimeError("받은 파일이 올바른 Excel 문서가 아닙니다. 공유 권한을 확인하세요.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        role = workbook_role(workbook.sheetnames)
        if role == "UNKNOWN":
            raise RuntimeError("기존 자료와 다른 시트 구조입니다. 시트 이름을 확인하세요.")
        return {"role": role, "sheets": list(workbook.sheetnames), "bytes": path.stat().st_size}
    finally:
        workbook.close()


def _publish_clean_cache(staging_dir: Path) -> tuple[Path, bool]:
    """Swap in a validated cache and retain the previous cache until commit."""
    cache_dir = GOOGLE_SHEETS_CACHE_DIR
    backup_dir = cache_dir.parent / f".{cache_dir.name}.backup-{uuid.uuid4().hex}"
    had_cache = cache_dir.exists()

    try:
        if had_cache:
            os.replace(cache_dir, backup_dir)
        os.replace(staging_dir, cache_dir)
    except Exception:
        if had_cache and backup_dir.exists() and not cache_dir.exists():
            os.replace(backup_dir, cache_dir)
        raise
    return backup_dir, had_cache


def _rollback_cache_swap(backup_dir: Path, had_cache: bool) -> None:
    if GOOGLE_SHEETS_CACHE_DIR.exists():
        shutil.rmtree(GOOGLE_SHEETS_CACHE_DIR)
    if had_cache and backup_dir.exists():
        os.replace(backup_dir, GOOGLE_SHEETS_CACHE_DIR)


def _finalize_cache_swap(backup_dir: Path) -> None:
    if backup_dir.exists():
        shutil.rmtree(backup_dir, ignore_errors=True)


def _point_source_records_to_cache(db_path: Path) -> None:
    """Replace temporary staging paths recorded by migration with durable cache paths."""
    with closing(connect(db_path)) as conn:
        for sheet in GOOGLE_SHEETS:
            conn.execute(
                "UPDATE source_files SET path=? WHERE file_name=?",
                (str(GOOGLE_SHEETS_CACHE_DIR / sheet["file_name"]), sheet["file_name"]),
            )
        conn.commit()


def sync_google_sheets(
    db_path: Path = DB_PATH,
    service_account_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Download the configured public Sheets as read-only XLSX files and merge them."""
    if not SYNC_LOCK.acquire(blocking=False):
        raise RuntimeError("다른 사용자가 이미 최신 자료를 반영하고 있습니다. 잠시 후 다시 시도하세요.")
    downloaded: list[dict[str, Any]] = []
    staging_dir: Path | None = None
    database_snapshot: Path | None = None
    database_migrated = False
    cache_swap: tuple[Path, bool] | None = None
    try:
        ensure_directories()
        staging_dir = Path(
            tempfile.mkdtemp(
                prefix=f".{GOOGLE_SHEETS_CACHE_DIR.name}.staging-",
                dir=GOOGLE_SHEETS_CACHE_DIR.parent,
            )
        )
        for sheet in GOOGLE_SHEETS:
            staged_path = staging_dir / sheet["file_name"]
            _download_xlsx(sheet["spreadsheet_id"], staged_path, service_account_info=service_account_info)
            details = _validate_workbook(staged_path)
            downloaded.append({"label": sheet["label"], "file": sheet["file_name"], **details})

        # Migrate only this clean, validated snapshot. The previous cache remains
        # untouched until the database migration has completed successfully.
        if db_path.exists():
            database_snapshot = db_path.with_name(f".{db_path.stem}.before-sync-{uuid.uuid4().hex}.db")
            with closing(connect(db_path)) as source_conn, closing(sqlite3.connect(database_snapshot)) as snapshot_conn:
                source_conn.backup(snapshot_conn)
        report = migrate(source_dir=staging_dir, db_path=db_path, reset=False)
        database_migrated = True
        cache_swap = _publish_clean_cache(staging_dir)
        _point_source_records_to_cache(db_path)
        synced_at = iso_now_kst()
        set_app_meta("last_google_sheets_sync_at", synced_at, db_path)
        set_app_meta("last_google_sheets_sync_status", "SUCCESS", db_path)
        _finalize_cache_swap(cache_swap[0])
        return {"synced_at": synced_at, "downloaded": downloaded, "report": report}
    except Exception as exc:
        if cache_swap is not None:
            _rollback_cache_swap(*cache_swap)
        if database_migrated and database_snapshot is not None and database_snapshot.exists():
            with closing(sqlite3.connect(database_snapshot)) as snapshot_conn, closing(connect(db_path)) as target_conn:
                snapshot_conn.backup(target_conn)
                target_conn.commit()
        set_app_meta("last_google_sheets_sync_status", f"ERROR: {exc}", db_path)
        raise
    finally:
        if staging_dir is not None and staging_dir.exists():
            shutil.rmtree(staging_dir)
        if database_snapshot is not None and database_snapshot.exists():
            database_snapshot.unlink()
        SYNC_LOCK.release()
