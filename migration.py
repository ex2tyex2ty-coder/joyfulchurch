from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sqlite3
import uuid
from collections import Counter, defaultdict
from contextlib import closing
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from config import DB_PATH, IMPORT_REPORT_PATH, QUALITY_IMPORTED, QUALITY_NEEDS_REVIEW, SOURCE_DIR, ensure_directories
from db import (
    canonical_service_key,
    canonicalize_service_records,
    connect,
    get_or_create_canonical_service,
    init_db,
    relink_attendance_events,
    series_key,
)
from time_utils import iso_now_kst, now_kst, today_kst


URL_RE = re.compile(r"https?://[^\s\]\)]+", re.I)
MONTHLY_RE = re.compile(r"^(20\d{2})_(\d{1,2})월$")
ATTENDANCE_RE = re.compile(r"^(20\d{2}) 예배인원$")
HONORIFICS = ("간사", "집사", "장로", "형제", "자매", "목사", "전도사")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def optional_count(value: Any) -> int | None:
    """Preserve a blank count as NULL and an explicit zero as 0."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        raise ValueError("참석 인원에는 true/false를 사용할 수 없습니다.")
    text = value.strip().replace(",", "") if isinstance(value, str) else str(value)
    try:
        numeric = Decimal(text)
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"정수 인원으로 읽을 수 없는 값: {value!r}") from exc
    if not numeric.is_finite() or numeric < 0 or numeric != numeric.to_integral_value():
        raise ValueError(f"참석 인원은 0 이상의 정수여야 합니다: {value!r}")
    # SQLite INTEGER is a signed 64-bit value. Reject an unsafe value here so
    # the source row is marked for review rather than failing the whole sync.
    if numeric > 9_223_372_036_854_775_807:
        raise ValueError(f"참석 인원이 저장 가능한 범위를 초과했습니다: {value!r}")
    return int(numeric)


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 40000 < value < 60000:
        try:
            result = from_excel(value)
            return result.date().isoformat() if isinstance(result, datetime) else result.isoformat()
        except Exception:
            return None
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def workbook_role(sheet_names: list[str]) -> str:
    names = set(sheet_names)
    if {"성찬식", "성인세례", "항시 체크 비품"}.issubset(names):
        return "MANUALS"
    attendance_years = {
        int(match.group(1))
        for name in names
        if (match := ATTENDANCE_RE.match(name))
    }
    lineup_years = {
        int(match.group(1))
        for name in names
        if (match := MONTHLY_RE.match(name))
    }
    if attendance_years & lineup_years:
        return "LINEUP_ATTENDANCE"
    return "UNKNOWN"


def sheet_rows(sheet, max_rows: int | None = None) -> list[list[Any]]:
    reported_max = sheet.max_row or max_rows or 1000
    limit = min(reported_max, max_rows or reported_max)
    return [list(row) for row in sheet.iter_rows(min_row=1, max_row=limit, values_only=True)]


def meaningful_dimensions(values: list[list[Any]]) -> tuple[int, int, int]:
    last_row = last_col = non_empty = 0
    for r, row_values in enumerate(values, start=1):
        for c, value in enumerate(row_values, start=1):
            if value not in (None, ""):
                last_row = max(last_row, r)
                last_col = max(last_col, c)
                non_empty += 1
    return last_row, last_col, non_empty


def insert_unresolved(
    conn: sqlite3.Connection,
    source_file_id: int,
    sheet_name: str,
    reason: str,
    raw_value: str = "",
    cell_reference: str = "",
    quality: str = QUALITY_NEEDS_REVIEW,
) -> None:
    conn.execute(
        "INSERT INTO unresolved_imports(source_file_id,sheet_name,cell_reference,raw_value,reason,quality) VALUES(?,?,?,?,?,?)",
        (source_file_id, sheet_name, cell_reference, raw_value[:2000], reason, quality),
    )


def add_reference(
    conn: sqlite3.Connection,
    title: str,
    url: str,
    source: str,
    manual_id: int | None = None,
    event_id: int | None = None,
    description: str = "",
) -> int:
    existing = conn.execute(
        "SELECT id FROM references_data WHERE url=? AND COALESCE(manual_id,0)=COALESCE(?,0) AND COALESCE(event_id,0)=COALESCE(?,0)",
        (url, manual_id, event_id),
    ).fetchone()
    if existing:
        conn.execute(
            "UPDATE references_data SET title=?,description=?,source=?,data_quality=?,archived_at=NULL WHERE id=?",
            (title[:160], description, source, QUALITY_IMPORTED, existing["id"]),
        )
        return int(existing["id"])
    cur = conn.execute(
        "INSERT INTO references_data(title,url,ref_type,description,manual_id,event_id,source,data_quality) VALUES(?,?,?,?,?,?,?,?)",
        (title[:160], url, "YouTube" if "youtube" in url or "youtu.be" in url else "웹 URL", description, manual_id, event_id, source, QUALITY_IMPORTED),
    )
    return int(cur.lastrowid)


def timing_offset(label: str, event_weekday: int | None) -> tuple[int | None, str]:
    """Return offset and quality. Monday=0; Sunday=6."""
    text = re.sub(r"\s+", "", label)
    if not text:
        return None, QUALITY_NEEDS_REVIEW
    if "차주주일" in text:
        return 7, QUALITY_IMPORTED
    if "1달" in text:
        return -30, QUALITY_IMPORTED
    if "2주전" in text:
        return -14, QUALITY_IMPORTED
    if "1주전주일" in text:
        return -7, QUALITY_IMPORTED
    if "1주전화요일" in text and event_weekday == 6:
        return -12, QUALITY_IMPORTED
    if "당일" in text or text in {"예배", "주일오전", "주일오후"}:
        return 0, QUALITY_IMPORTED
    weekdays = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "주일": 6}
    if event_weekday is not None:
        for prefix, weekday in weekdays.items():
            if text.startswith(prefix):
                delta = weekday - event_weekday
                if delta > 0:
                    delta -= 7
                return delta, QUALITY_IMPORTED
    return None, QUALITY_NEEDS_REVIEW


def markdown_from_rows(values: list[list[Any]], max_rows: int = 100) -> str:
    lines: list[str] = []
    for row_number, row_values in enumerate(values[:max_rows], start=1):
        cells = [clean(value) for value in row_values]
        if not any(cells):
            continue
        first = next((cell for cell in cells if cell), f"{row_number}행")
        remaining = cells[cells.index(first) + 1:]
        lines.append(f"- **{first}**")
        for value in remaining:
            if value:
                lines.append("  - " + value.replace("\n", "  \n    "))
    return "\n".join(lines)


def task_header_index(values: list[list[Any]]) -> int | None:
    for index, row_values in enumerate(values[:5]):
        headers = [re.sub(r"\s+", "", clean(value)) for value in row_values]
        if "준비물" in headers and any("준비내용" in header for header in headers):
            return index
    return None


def is_timing_label(value: str) -> bool:
    text = re.sub(r"\s+", "", value)
    if not text or len(text) > 40:
        return False
    return bool(
        re.search(r"(\d+달|\d+주전|차주|당일|예배|주일|오전|오후|월요일|화요일|수요일|목요일|금요일|토요일)", text)
        or text in {"월", "화", "수", "목", "금", "토"}
    )


def extract_task_records(
    values: list[list[Any]], header_index: int, last_row: int, event_weekday: int | None
) -> tuple[list[dict[str, Any]], list[str]]:
    records: list[dict[str, Any]] = []
    guidance: list[str] = []
    timing = ""
    for row_number, row_values in enumerate(values[header_index + 1:last_row], start=header_index + 2):
        cells = list(row_values) + [None] * max(0, 7 - len(row_values))
        first = clean(cells[0])
        item = clean(cells[1])
        quantity = clean(cells[2])
        task_text = clean(cells[3])
        note = clean(cells[4])
        extras = [clean(value) for value in cells[5:] if clean(value)]

        if first and "업데이트" not in first:
            if is_timing_label(first):
                timing = first
            elif not item and not task_text:
                supplemental = "\n".join(filter(None, [first, note, *extras]))
                if supplemental:
                    guidance.append(supplemental.lstrip("*- \t"))

        if not item and not task_text:
            continue
        if item == "준비물" or re.sub(r"\s+", "", task_text) == "준비내용":
            continue

        urls = extract_urls([item, task_text, note, *extras])
        non_url_text = " ".join(filter(None, [item, task_text, note, *extras]))
        if urls and not URL_RE.sub("", non_url_text).strip(" ·|/-"):
            continue

        title = next((line.strip() for line in task_text.splitlines() if line.strip()), item)
        description_parts: list[str] = []
        if item:
            description_parts.append(f"- 준비물: {item}")
        if quantity:
            description_parts.append(f"- 수량: {quantity}")
        task_lines = [line.strip() for line in task_text.splitlines() if line.strip()]
        if len(task_lines) > 1:
            description_parts.append("- 작업 상세:\n  " + "\n  ".join(task_lines[1:]))
        if note:
            description_parts.append(f"- 비고: {note}")
        if extras:
            description_parts.append("- 추가 정보:\n  " + "\n  ".join(extras))
        offset, quality = timing_offset(timing, event_weekday)
        records.append({
            "row_number": row_number,
            "timing": timing,
            "item": item,
            "quantity": quantity,
            "title": title[:240],
            "description": "\n".join(description_parts),
            "offset": offset,
            "quality": quality,
        })
    return records, guidance


def structured_task_markdown(records: list[dict[str, Any]], guidance: list[str]) -> str:
    lines = ["## 준비 일정"]
    current_timing = None
    for record in records:
        timing = record["timing"] or "준비 시점 미기록"
        if timing != current_timing:
            lines.append(f"\n### {timing}")
            current_timing = timing
        lines.append(f"- **{record['title']}**")
        if record["description"]:
            lines.extend(f"  {line}" for line in record["description"].splitlines())
    if guidance:
        lines.append("\n## 추가 운영 안내")
        for item in guidance:
            lines.append("- " + item.replace("\n", "  \n  "))
    return "\n".join(lines)


def latest_update_label(values: list[list[Any]]) -> str | None:
    for row_values in values:
        for value in row_values:
            text = clean(value)
            if "업데이트" not in text:
                continue
            matched = re.search(r"(?<!\d)(\d{2,4})[.\-/]\s*(\d{1,2})[.\-/]\s*(\d{1,2})", text)
            if matched:
                year, month, day = map(int, matched.groups())
                if year < 100:
                    year += 2000
                try:
                    return date(year, month, day).isoformat()
                except ValueError:
                    return text
            return text
    return None


def create_manual(
    conn: sqlite3.Connection,
    title: str,
    category: str,
    source_file: str,
    source_sheet: str,
    how_text: str,
    current_standard: str,
    quality: str = QUALITY_IMPORTED,
) -> int:
    existing = conn.execute("SELECT id,data_quality FROM manuals WHERE source=? AND source_sheet=?", (source_file, source_sheet)).fetchone()
    if existing:
        manual_id = int(existing["id"])
        current = conn.execute(
            "SELECT * FROM manual_revisions WHERE manual_id=? AND status='CURRENT' ORDER BY version DESC LIMIT 1",
            (manual_id,),
        ).fetchone()
        source_revision_summaries = {"기존 Spreadsheet 최초 이관", "Google Sheets 자동 동기화"}
        user_current = bool(
            current
            and existing["data_quality"] == "Verified"
            and str(current["change_summary"] or "") not in source_revision_summaries
        )
        if user_current:
            conn.execute(
                "UPDATE manuals SET title=?,category=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (title, category, manual_id),
            )
        else:
            conn.execute(
                "UPDATE manuals SET title=?,category=?,current_standard=?,data_quality=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (title, category, current_standard, quality, manual_id),
            )
        if not user_current and (not current or (current["how_text"] or "").strip() != how_text.strip()):
            next_version = int(conn.execute(
                "SELECT COALESCE(MAX(version),0)+1 AS version FROM manual_revisions WHERE manual_id=?", (manual_id,)
            ).fetchone()["version"])
            conn.execute("UPDATE manual_revisions SET status='SUPERSEDED' WHERE manual_id=? AND status='CURRENT'", (manual_id,))
            conn.execute(
                "INSERT INTO manual_revisions(manual_id,version,what_text,how_text,why_text,caution,change_summary,status) "
                "VALUES(?,?,?,?,?,?,?,'CURRENT')",
                (manual_id, next_version, f"{source_sheet} 운영 기준", how_text,
                 "Google 스프레드시트 원본에서 동기화했습니다.",
                 "원본 시트 변경사항을 확인한 뒤 운영에 적용하세요.", "Google Sheets 자동 동기화"),
            )
            conn.execute("UPDATE manuals SET version=? WHERE id=?", (next_version, manual_id))
        return manual_id
    cur = conn.execute(
        "INSERT INTO manuals(title,category,current_standard,source,source_sheet,data_quality) VALUES(?,?,?,?,?,?)",
        (title, category, current_standard, source_file, source_sheet, quality),
    )
    manual_id = int(cur.lastrowid)
    conn.execute(
        "INSERT INTO manual_revisions(manual_id,version,what_text,how_text,why_text,caution,change_summary,status) VALUES(?,?,?,?,?,?,?,'CURRENT')",
        (manual_id, 1, f"{source_sheet} 운영 기준", how_text, "원본 자료에는 결정 이유가 별도로 기록되어 있지 않습니다.",
         "원본 의미가 불명확한 항목은 Needs Review 상태로 보존했습니다.", "기존 Spreadsheet 최초 이관"),
    )
    return manual_id


def upsert_task_template(
    conn: sqlite3.Connection,
    event_template_id: int,
    title: str,
    description: str,
    source_timing: str,
    due_offset: int | None,
    source: str,
    quality: str,
) -> int:
    existing = conn.execute("SELECT id FROM task_templates WHERE source=? ORDER BY id LIMIT 1", (source,)).fetchone()
    if existing:
        conn.execute(
            "UPDATE task_templates SET event_template_id=?,title=?,description=?,source_timing=?,due_offset=?,data_quality=? WHERE id=?",
            (event_template_id, title, description, source_timing, due_offset, quality, existing["id"]),
        )
        return int(existing["id"])
    cur = conn.execute(
        "INSERT INTO task_templates(event_template_id,title,description,source_timing,due_offset,source,data_quality) VALUES(?,?,?,?,?,?,?)",
        (event_template_id, title, description, source_timing, due_offset, source, quality),
    )
    return int(cur.lastrowid)


def extract_urls(texts: Iterable[Any]) -> list[str]:
    result: list[str] = []
    for value in texts:
        result.extend(URL_RE.findall(clean(value)))
    return list(dict.fromkeys(result))


def import_manual_workbook(conn: sqlite3.Connection, path: Path, source_file_id: int, report: dict[str, Any]) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    credential_sheets = {"나스"}
    event_exclusions = {"항시 체크 비품", "특별예배 관련 내용", "선교 파송보고예배", "시트17", "시트18"}
    manual_count = template_count = task_template_count = reference_count = 0

    for sheet in wb.worksheets:
        values = sheet_rows(sheet, max_rows=min(sheet.max_row or 150, 150))
        last_row, last_col, non_empty = meaningful_dimensions(values)
        if sheet.title in credential_sheets:
            insert_unresolved(
                conn, source_file_id, sheet.title,
                "계정·비밀번호가 평문으로 저장된 보안 민감 시트. 값은 DB에 가져오지 않았으며 즉시 비밀번호 변경이 필요합니다.",
                raw_value="[REDACTED]",
            )
            continue
        if not non_empty:
            continue

        header_index = task_header_index(values)
        event_weekday = 4 if "성금요일" in sheet.title else (None if sheet.title == "수련회" else 6)
        task_records: list[dict[str, Any]] = []
        guidance: list[str] = []
        if header_index is not None:
            task_records, guidance = extract_task_records(values, header_index, last_row, event_weekday)
            how_text = structured_task_markdown(task_records, guidance)
            update_label = latest_update_label(values)
            standard = f"준비업무 {len(task_records)}건" + (f" · 최종 표기 {update_label}" if update_label else " · 원본 엑셀 기준")
        else:
            how_text = markdown_from_rows(values[:last_row], max_rows=150)
            meaningful_rows = sum(1 for row_values in values[:last_row] if any(clean(value) for value in row_values[:last_col]))
            standard_overrides = {
                "항시 체크 비품": f"상시 비품 점검 {max(meaningful_rows - 1, 0)}건",
                "특별예배 관련 내용": "특별예배 유형 4개 운영기준",
                "선교 파송보고예배": f"선교 예배 기록 {max(meaningful_rows - 1, 0)}건",
                "시트17": f"카메라 운영지침 {max(meaningful_rows - 1, 0)}건",
                "시트18": f"장소 3곳 비교 · 항목 {max(meaningful_rows - 1, 0)}개",
            }
            standard = standard_overrides.get(sheet.title, f"원본 운영자료 {meaningful_rows}개 행")
        category = "특별예배" if sheet.title not in {"항시 체크 비품", "시트17", "시트18"} else "상시운영"
        display_title = {"시트17": "카메라 운영 가이드", "시트18": "수련회 장소 비교표"}.get(sheet.title, f"{sheet.title} 운영 매뉴얼")
        manual_id = create_manual(
            conn, display_title, category, path.name, sheet.title,
            how_text, standard, QUALITY_IMPORTED,
        )
        manual_count += 1

        conn.execute(
            "UPDATE references_data SET archived_at=CURRENT_TIMESTAMP WHERE manual_id=? AND source LIKE ?",
            (manual_id, f"{path.name} / {sheet.title}%"),
        )

        for r, row_values in enumerate(values[:last_row], start=1):
            for url in extract_urls(row_values[:last_col]):
                context_parts = [URL_RE.sub("", clean(value)).strip(" ·|/-") for value in row_values[:last_col]]
                context = " · ".join(value for value in context_parts if value)
                add_reference(
                    conn, f"{sheet.title} 참고자료 · {r}행", url, f"{path.name} / {sheet.title}!{r}",
                    manual_id=manual_id, description=context[:1000],
                )
                reference_count += 1

        if sheet.title not in event_exclusions and header_index is not None:
            cur = conn.execute(
                "INSERT OR IGNORE INTO event_templates(title,category,description,manual_id,source,source_sheet,data_quality) VALUES(?,?,?,?,?,?,?)",
                (sheet.title, "특별예배", f"{sheet.title} 원본 매뉴얼 기반 템플릿", manual_id, path.name, sheet.title, QUALITY_IMPORTED),
            )
            template = conn.execute("SELECT id FROM event_templates WHERE title=?", (sheet.title,)).fetchone()
            if cur.rowcount:
                template_count += 1
            event_template_id = int(template["id"])
            conn.execute(
                "UPDATE task_templates SET data_quality='Stale' WHERE event_template_id=? AND source LIKE ?",
                (event_template_id, f"{path.name} / {sheet.title}!%"),
            )
            for task_record in task_records:
                upsert_task_template(
                    conn, event_template_id, task_record["title"], task_record["description"], task_record["timing"], task_record["offset"],
                    f"{path.name} / {sheet.title}!{task_record['row_number']}", task_record["quality"],
                )
                task_template_count += 1
                if task_record["offset"] is None and task_record["timing"]:
                    insert_unresolved(
                        conn, source_file_id, sheet.title, "행사일 기준 날짜로 자동 변환하기 어려운 준비시점",
                        task_record["timing"], f"A{task_record['row_number']}",
                    )

        if (sheet.max_row or 0) >= 900 and last_row < 100:
            insert_unresolved(
                conn, source_file_id, sheet.title,
                f"실제 데이터는 {last_row}행까지이나 사용 범위가 {sheet.max_row}행으로 과도하게 잡혀 있음",
                quality="Imported",
            )

    report["imported"]["manuals"] += manual_count
    report["imported"]["event_templates"] += template_count
    report["imported"]["task_templates"] += task_template_count
    report["imported"]["references"] += reference_count
    wb.close()


def find_date_row(values: list[list[Any]]) -> tuple[int, list[int]] | None:
    for row_index, row_values in enumerate(values[:15]):
        date_cols = [col for col, value in enumerate(row_values) if iso_date(value)]
        if len(date_cols) >= 2:
            return row_index, date_cols
    return None


def role_label(row_values: list[Any], first_date_col: int) -> str:
    for value in row_values[:first_date_col]:
        text = clean(value).replace("\n", "").replace(" ", "")
        if text:
            return text
    return ""


def normalize_person(raw: str) -> str:
    name = raw.strip()
    for honorific in HONORIFICS:
        name = name.replace(honorific, "")
    return re.sub(r"\s+", "", name)


def people(raw: str) -> list[str]:
    if not raw or raw.strip() in {"-", "없음"}:
        return []
    tokens = re.split(r"[/,&·]|\s+&\s+", raw)
    result = [normalize_person(token) for token in tokens]
    return [name for name in result if name and len(name) <= 12 and not name.isdigit()]


ROLE_MAP = {
    "대표기도": "대표기도", "음향": "음향", "조명/카메라": "조명/카메라", "FD": "FD", "자막/송출": "자막/송출",
    "하이라이트촬영": "하이라이트촬영", "메인촬영": "메인촬영", "예배세팅최종담당자": "예배세팅 최종", "뒷정리최종담당자": "뒷정리 최종",
}


def match_template(conn: sqlite3.Connection, special: str) -> int | None:
    compact = re.sub(r"\s+", "", special)
    templates = conn.execute("SELECT id,title FROM event_templates WHERE status='CURRENT'").fetchall()
    for template in templates:
        candidate = re.sub(r"\s+", "", template["title"])
        if candidate in compact or compact in candidate:
            return int(template["id"])
    return None


def import_lineup_workbook(conn: sqlite3.Connection, path: Path, source_file_id: int, report: dict[str, Any]) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    imported_months: set[tuple[int, int]] = set()
    service_count = assignment_count = member_count = event_count = attendance_count = manual_count = task_template_count = event_template_count = 0

    # Attendance is imported first so the actual source rows are preserved.
    for sheet in wb.worksheets:
        match = ATTENDANCE_RE.match(sheet.title)
        if not match:
            continue
        conn.execute("DELETE FROM attendance WHERE source=? AND source_sheet=?", (path.name, sheet.title))
        source_year = int(match.group(1))
        values = sheet_rows(sheet, max_rows=min(sheet.max_row or 500, 500))
        for row_number, row_values in enumerate(values[1:], start=2):
            cells = list(row_values) + [None] * 7
            service_date = iso_date(cells[0])
            service_type = clean(cells[1])
            if not service_date or not service_type:
                continue
            raw_online = cells[2]
            raw_offline = cells[3]
            raw_total = cells[4]
            invalid_counts: list[str] = []
            try:
                online = optional_count(raw_online)
            except (TypeError, ValueError):
                online = None
                invalid_counts.append(f"온라인={raw_online!r}")
            try:
                offline = optional_count(raw_offline)
            except (TypeError, ValueError):
                offline = None
                invalid_counts.append(f"현장={raw_offline!r}")
            try:
                original_total = optional_count(raw_total)
            except (TypeError, ValueError):
                original_total = None
                invalid_counts.append(f"총계={raw_total!r}")
            components_present = raw_online is not None or raw_offline is not None
            calculated = int(online or 0) + int(offline or 0)
            stored_total = calculated if components_present else original_total
            if max(int(online or 0), int(offline or 0), int(original_total or 0)) > 0:
                record_status = "COUNTED"
            elif raw_online is None and raw_offline is None:
                record_status = "PENDING"
            else:
                # An explicit 0 is not silently treated as a completed count.
                record_status = "UNKNOWN"
            quality = QUALITY_IMPORTED
            reasons: list[str] = []
            if date.fromisoformat(service_date).year != source_year:
                quality = QUALITY_NEEDS_REVIEW
                reasons.append(f"{source_year} 시트에 {service_date} 날짜가 포함됨")
            if original_total is not None and components_present and original_total != calculated:
                quality = QUALITY_NEEDS_REVIEW
                reasons.append(f"원본 총계 {original_total}와 온라인+오프라인 {calculated} 불일치")
            if record_status == "PENDING":
                quality = QUALITY_NEEDS_REVIEW
                reasons.append("날짜와 예배구분은 있으나 참석 인원이 미입력 상태")
            elif record_status == "UNKNOWN":
                quality = QUALITY_NEEDS_REVIEW
                reasons.append("온라인·현장에 0이 명시되어 집계 완료 여부 확인 필요")
            if invalid_counts:
                quality = QUALITY_NEEDS_REVIEW
                record_status = "UNKNOWN"
                reasons.append("숫자로 읽을 수 없는 원본 값: " + ", ".join(invalid_counts))
            if online is not None and offline is not None and online > offline:
                quality = QUALITY_NEEDS_REVIEW
                reasons.append("온라인 인원이 오프라인보다 커 열 입력 순서 확인 필요")
            service_exists = conn.execute(
                "SELECT id FROM services WHERE canonical_key=?",
                (canonical_service_key(service_date, service_type),),
            ).fetchone()
            service_id = get_or_create_canonical_service(
                conn,
                service_date,
                service_type,
                source=path.name,
                source_sheet=sheet.title,
                data_quality=quality,
            )
            if service_exists is None:
                service_count += 1
            conn.execute(
                "INSERT INTO attendance(service_date,service_type,online_count,offline_count,total_count,record_status,"
                "raw_online_count,raw_offline_count,raw_total_count,metric_type,measurement_note,service_id,notes,source,source_sheet,source_row,data_quality) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    service_date,
                    service_type,
                    online,
                    offline,
                    stored_total,
                    record_status,
                    online,
                    offline,
                    original_total,
                    "ONSITE_PLUS_ONLINE_UNDEFINED",
                    "원본 시트의 온라인·현장·총계 열을 보존함; 온라인 집계 기준은 원본에 명시되지 않음",
                    service_id,
                    "; ".join(reasons),
                    path.name,
                    sheet.title,
                    row_number,
                    quality,
                ),
            )
            attendance_count += 1
            if reasons:
                insert_unresolved(conn, source_file_id, sheet.title, "; ".join(reasons), f"{service_date} / {service_type}", f"A{row_number}:E{row_number}")

    for sheet in wb.worksheets:
        match = MONTHLY_RE.match(sheet.title)
        if not match:
            continue
        year, month = map(int, match.groups())
        imported_months.add((year, month))
        conn.execute("DELETE FROM assignments WHERE source LIKE ?", (f"{path.name} / {sheet.title}%",))
        values = sheet_rows(sheet, max_rows=min(sheet.max_row or 120, 120))
        found = find_date_row(values)
        if not found:
            insert_unresolved(conn, source_file_id, sheet.title, "월별 라인업에서 날짜 행을 찾지 못함")
            continue
        date_row, date_cols = found
        first_date_col = min(date_cols)
        roles: dict[str, int] = {}
        special_row = video_row = None
        for row_index in range(date_row + 1, min(len(values), date_row + 20)):
            label = role_label(values[row_index], first_date_col)
            if label in {"특별순서", "특순"}:
                special_row = row_index
            elif label == "영상":
                video_row = row_index
            for source_label in ROLE_MAP:
                if label == source_label:
                    roles[ROLE_MAP[source_label]] = row_index
        for col_index in date_cols:
            service_date = iso_date(values[date_row][col_index])
            if not service_date:
                continue
            weekday = clean(values[date_row - 1][col_index]) if date_row > 0 else ""
            service_type = "주일예배" if "주일" in weekday else ("금요예배" if "금" in weekday else weekday or "예배")
            special = clean(values[special_row][col_index]) if special_row is not None and col_index < len(values[special_row]) else ""
            prayer = clean(values[roles["대표기도"]][col_index]) if "대표기도" in roles and col_index < len(values[roles["대표기도"]]) else ""
            quality = QUALITY_IMPORTED
            if date.fromisoformat(service_date).month != month:
                quality = QUALITY_NEEDS_REVIEW
                insert_unresolved(conn, source_file_id, sheet.title, f"시트 월({month}월)과 날짜({service_date})가 일치하지 않음", service_date)
            service_exists = conn.execute(
                "SELECT id FROM services WHERE canonical_key=?",
                (canonical_service_key(service_date, service_type),),
            ).fetchone()
            service_id = get_or_create_canonical_service(
                conn,
                service_date,
                service_type,
                special_sequence=special,
                representative_prayer=prayer,
                source=path.name,
                source_sheet=sheet.title,
                data_quality=quality,
            )
            if service_exists is None:
                service_count += 1

            for role, role_row in roles.items():
                raw = clean(values[role_row][col_index]) if col_index < len(values[role_row]) else ""
                for name in people(raw):
                    member_cur = conn.execute(
                        "INSERT OR IGNORE INTO members(name,team,role,source,data_quality) VALUES(?,?,?,?,?)",
                        (name, "예배팀", role, f"{path.name} / {sheet.title}", QUALITY_IMPORTED),
                    )
                    if member_cur.rowcount:
                        member_count += 1
                    member = conn.execute("SELECT id FROM members WHERE name=?", (name,)).fetchone()
                    assignment_cur = conn.execute(
                        "INSERT OR IGNORE INTO assignments(service_id,member_id,role,raw_name,source,data_quality) VALUES(?,?,?,?,?,?)",
                        (service_id, member["id"], role, raw, f"{path.name} / {sheet.title}", quality),
                    )
                    if assignment_cur.rowcount:
                        assignment_count += 1

            if special and special not in {"-", "없음"}:
                template_id = match_template(conn, special)
                event_title = f"{date.fromisoformat(service_date).year} {special.replace(chr(10), ' / ')}"
                existing = conn.execute("SELECT id FROM events WHERE event_date=? AND title=?", (service_date, event_title)).fetchone()
                if existing:
                    event_id = int(existing["id"])
                    imported_status = "COMPLETED" if service_date < today_kst().isoformat() else "PLANNING"
                    conn.execute(
                        "UPDATE events SET service_id=?,service_type=?,"
                        "status=CASE WHEN source IS NOT NULL THEN ? ELSE status END,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (service_id, service_type, imported_status, event_id),
                    )
                else:
                    previous = conn.execute(
                        "SELECT id FROM events WHERE series_key=? AND event_date < ? ORDER BY event_date DESC LIMIT 1",
                        (series_key(special), service_date),
                    ).fetchone()
                    event_cur = conn.execute(
                        "INSERT INTO events(title,series_key,category,event_date,status,event_template_id,previous_event_id,service_id,service_type,source,source_event_id,data_quality) "
                        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                        (event_title, series_key(special), "특별순서", service_date, "COMPLETED" if service_date < today_kst().isoformat() else "PLANNING",
                         template_id, previous["id"] if previous else None, service_id, service_type, path.name, f"{sheet.title}:{col_index + 1}", quality),
                    )
                    event_id = int(event_cur.lastrowid)
                    event_count += 1
                    if template_id:
                        templates = conn.execute(
                            "SELECT * FROM task_templates WHERE event_template_id=? AND data_quality<>'Stale'",
                            (template_id,),
                        ).fetchall()
                        for task in templates:
                            due_date = None
                            if task["due_offset"] is not None:
                                due_date = (date.fromisoformat(service_date) + timedelta(days=int(task["due_offset"]))).isoformat()
                            conn.execute(
                                "INSERT INTO tasks(event_id,task_template_id,title,description,owner,status,priority,source_timing,due_offset,due_date,source,data_quality) VALUES(?,?,?,?,?,'TODO',?,?,?,?,?,?)",
                                (event_id, task["id"], task["title"], task["description"], task["default_owner"], task["priority"], task["source_timing"],
                                 task["due_offset"], due_date, task["source"], task["data_quality"]),
                            )
                # Never use a date-only update: another worship type may share this date.
                conn.execute(
                    "UPDATE attendance SET event_id=? WHERE service_id=?",
                    (event_id, service_id),
                )
                if video_row is not None and col_index < len(values[video_row]):
                    for url in extract_urls([values[video_row][col_index]]):
                        add_reference(conn, f"{special} 참고 영상", url, f"{path.name} / {sheet.title}", event_id=event_id)

    # Current regular-service setup timetables become manuals and task templates.
    for sheet_name, template_title, weekday in [
        ("주일 세팅 타임테이블", "주일예배", 6),
        ("금요집회 세팅 타임테이블", "금요집회", 4),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        values = sheet_rows(sheet, max_rows=min(sheet.max_row or 100, 100))
        manual_id = create_manual(conn, f"{sheet_name} 운영 매뉴얼", "정기예배", path.name, sheet_name, markdown_from_rows(values), sheet_name)
        manual_count += 1
        template_cur = conn.execute(
            "INSERT OR IGNORE INTO event_templates(title,category,description,manual_id,source,source_sheet,data_quality) VALUES(?,?,?,?,?,?,?)",
            (template_title, "정기예배", f"{sheet_name} 기반", manual_id, path.name, sheet_name, QUALITY_IMPORTED),
        )
        if template_cur.rowcount:
            event_template_count += 1
        template_id = int(conn.execute("SELECT id FROM event_templates WHERE title=?", (template_title,)).fetchone()["id"])
        conn.execute(
            "UPDATE task_templates SET data_quality='Stale' WHERE event_template_id=? AND source LIKE ?",
            (template_id, f"{path.name} / {sheet_name}!%"),
        )
        timing = section = ""
        for row_number, row_values in enumerate(values[1:], start=2):
            cells = list(row_values) + [None] * 5
            if cells[0] not in (None, ""):
                timing = clean(cells[0])
            if clean(cells[1]):
                section = clean(cells[1])
            task = clean(cells[2])
            if not task:
                continue
            title = task.splitlines()[0][:240]
            description = task + (f"\n준비사항: {clean(cells[3])}" if clean(cells[3]) else "")
            upsert_task_template(
                conn, template_id, title, description, f"{timing} {section}".strip(), 0,
                f"{path.name} / {sheet_name}!{row_number}", QUALITY_IMPORTED,
            )
            task_template_count += 1

    # Known structural quality observations.
    years = sorted({year for year, _ in imported_months})
    for year in years:
        if (year, 10) not in imported_months:
            insert_unresolved(conn, source_file_id, "월별 라인업", f"{year}년 10월 라인업 시트가 없음")
    if "2025_11월의 사본" in wb.sheetnames:
        insert_unresolved(conn, source_file_id, "2025_11월의 사본", "원본 2025_11월 시트와 잠재적 중복. 자동 Import에서 제외", quality="Imported")

    report["imported"]["attendance"] += attendance_count
    report["imported"]["services"] += service_count
    report["imported"]["assignments"] += assignment_count
    report["imported"]["members"] += member_count
    report["imported"]["events"] += event_count
    report["imported"]["manuals"] += manual_count
    report["imported"]["event_templates"] += event_template_count
    report["imported"]["task_templates"] += task_template_count
    wb.close()


def migrate(
    source_dir: Path = SOURCE_DIR,
    db_path: Path = DB_PATH,
    reset: bool = False,
    report_path: Path | None = IMPORT_REPORT_PATH,
) -> dict[str, Any]:
    ensure_directories()
    source_files = sorted(
        path for path in source_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}
    )
    if reset:
        if not source_files:
            raise RuntimeError("재구축할 Excel 원본이 없습니다. 최신 자료 업데이트를 먼저 완료하세요.")
        temporary_db = db_path.with_name(f".{db_path.stem}.rebuild-{uuid.uuid4().hex}.db")
        try:
            report = migrate(
                source_dir=source_dir,
                db_path=temporary_db,
                reset=False,
                report_path=None,
            )
            with closing(connect(temporary_db)) as validation_conn:
                integrity = validation_conn.execute("PRAGMA quick_check").fetchone()[0]
                foreign_keys = validation_conn.execute("PRAGMA foreign_key_check").fetchall()
            if integrity != "ok" or foreign_keys:
                raise RuntimeError("새 데이터베이스 무결성 검사를 통과하지 못했습니다.")
            if db_path.exists():
                backup_dir = db_path.parent / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                backup_path = backup_dir / f"{db_path.stem}_before_rebuild_{now_kst().strftime('%Y%m%d_%H%M%S')}.db"
                with closing(connect(db_path)) as source_conn, closing(sqlite3.connect(backup_path)) as backup_conn:
                    source_conn.backup(backup_conn)
                    source_conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
                for sidecar in (
                    db_path.with_name(db_path.name + "-wal"),
                    db_path.with_name(db_path.name + "-shm"),
                ):
                    if sidecar.exists():
                        sidecar.unlink()
            os.replace(temporary_db, db_path)
            if report_path is not None:
                report_path.write_text(
                    json.dumps(report, ensure_ascii=False, indent=2, default=str),
                    encoding="utf-8",
                )
            return report
        finally:
            for leftover in (
                temporary_db,
                temporary_db.with_name(temporary_db.name + "-wal"),
                temporary_db.with_name(temporary_db.name + "-shm"),
            ):
                if leftover.exists():
                    leftover.unlink()
    init_db(db_path)
    report: dict[str, Any] = {
        "started_at": iso_now_kst(),
        "source_directory": str(source_dir),
        "source_files": [],
        "imported": defaultdict(int),
        "skipped": [],
        "unresolved": 0,
        "data_quality": {},
    }
    files = source_files
    with closing(connect(db_path)) as conn:
        run_cur = conn.execute("INSERT INTO import_runs(status) VALUES('RUNNING')")
        run_id = int(run_cur.lastrowid)
        for path in files:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = []
            for sheet in wb.worksheets:
                values = sheet_rows(sheet, max_rows=min(sheet.max_row or 150, 150))
                last_row, last_col, non_empty = meaningful_dimensions(values)
                sheets.append({
                    "name": sheet.title,
                    "reported_max_row": sheet.max_row,
                    "reported_max_column": sheet.max_column,
                    "meaningful_last_row": last_row,
                    "meaningful_last_column": last_col,
                    "non_empty_cells": non_empty,
                })
            role = workbook_role(wb.sheetnames)
            wb.close()
            source_file = conn.execute("SELECT id FROM source_files WHERE file_name=? ORDER BY id LIMIT 1", (path.name,)).fetchone()
            if source_file:
                conn.execute(
                    "UPDATE source_files SET path=?,file_size=?,sha256=?,workbook_role=?,sheet_inventory_json=?,"
                    "imported_at=CURRENT_TIMESTAMP WHERE id=?",
                    (str(path), path.stat().st_size, file_hash(path), role, json.dumps(sheets, ensure_ascii=False), source_file["id"]),
                )
            else:
                cur = conn.execute(
                    "INSERT INTO source_files(path,file_name,file_size,sha256,workbook_role,sheet_inventory_json) VALUES(?,?,?,?,?,?)",
                    (str(path), path.name, path.stat().st_size, file_hash(path), role, json.dumps(sheets, ensure_ascii=False)),
                )
                source_file = {"id": int(cur.lastrowid)}
            source_file_id = int(source_file["id"])
            conn.execute("DELETE FROM unresolved_imports WHERE source_file_id=?", (source_file_id,))
            report["source_files"].append({"file": path.name, "role": role, "size": path.stat().st_size, "sheets": sheets})
            if role == "MANUALS":
                import_manual_workbook(conn, path, source_file_id, report)
            elif role == "LINEUP_ATTENDANCE":
                import_lineup_workbook(conn, path, source_file_id, report)
            else:
                report["skipped"].append({"file": path.name, "reason": "알려진 Spreadsheet 구조와 일치하지 않음"})
                insert_unresolved(conn, source_file_id, "", "Workbook 역할을 자동 판별하지 못함", path.name)

        # A sync may contain attendance and lineup sheets in either order. Resolve
        # every consumer to the same canonical occurrence before publishing it.
        canonicalize_service_records(conn)
        relink_attendance_events(conn)
        report["unresolved"] = int(conn.execute("SELECT COUNT(*) AS count FROM unresolved_imports").fetchone()["count"])
        quality_rows = conn.execute("SELECT data_quality, COUNT(*) AS count FROM attendance GROUP BY data_quality").fetchall()
        report["data_quality"] = {item["data_quality"]: item["count"] for item in quality_rows}
        report["imported"] = dict(report["imported"])
        report["finished_at"] = iso_now_kst()
        report_json = json.dumps(report, ensure_ascii=False, indent=2, default=str)
        conn.execute("UPDATE import_runs SET finished_at=CURRENT_TIMESTAMP,status='SUCCESS',report_json=? WHERE id=?", (report_json, run_id))
        conn.execute(
            "INSERT INTO app_meta(key,value) VALUES('last_import_at',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=CURRENT_TIMESTAMP",
            (report["finished_at"],),
        )
        conn.commit()
    if report_path is not None:
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="JOYFUL WORSHIP OPS Spreadsheet migration")
    parser.add_argument("--source", type=Path, default=SOURCE_DIR)
    parser.add_argument("--db", type=Path, default=DB_PATH)
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()
    report = migrate(args.source, args.db, args.reset)
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
